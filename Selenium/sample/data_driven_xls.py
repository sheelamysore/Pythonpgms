from selenium import webdriver
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
import time 
import openpyxl

opt=ChromeOptions()
opt.add_experimental_option("detach", True)

path=r"C:\Users\admin\eclipse-workspace\ORG_HRM_LoginData.xlsx"
my_workbook=openpyxl.load_workbook(path)
active_sheet=my_workbook.active
rows=active_sheet.max_row
col=active_sheet.max_column
# print(rows,col)
for i in range (2,rows+1):
    username=active_sheet.cell(row=i,column=1).value
    print(username)
    password=active_sheet.cell(row=i,column=2).value
    print(password)
   # for j in range(1,col+1):
    driver=webdriver.Chrome(options=opt)
    driver.maximize_window()

    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")    
    driver.implicitly_wait(10)      
    username_textbox=driver.find_element(By.XPATH,"((//input)[2])")
    username_textbox.send_keys(username)
    print("username entered",username)
    username_textbox1=driver.find_element(By.XPATH,"(//input)[3]")
    username_textbox1.send_keys(password)
    print("password entered",password)
    login_button=driver.find_element(By.XPATH,"//button")
    login_button.click()
    page_url=driver.current_url
    if "/dashboard/index" in page_url:
        print("Passed")
    
    else:
        print("Failed")
    
    
    
    
    
    